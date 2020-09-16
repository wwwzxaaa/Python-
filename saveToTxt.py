__author__ = 'Lin'
#coding=utf-8

import openpyxl
def excel_to_txt(path_file, save_file_name):
    """
    将excel中的保存到txt中
    :param path_file: excel文件路径
    :param save_file_name: txt文件名
    :return:
    """
    wb = openpyxl.load_workbook(path_file)
    current_sheet = wb.active
    maxRow = current_sheet.max_row
    maxColumn = current_sheet.max_column
    if maxRow != maxColumn:
        print("数据存在问题:最大行数和最大列数的数值不相等")
        return

    diff_num = 0
    #对称位置遍历
    fd = open(save_file_name + '.txt', 'w', encoding='utf-8')

    for i in range(2, maxRow+1):
        i_pos = current_sheet.cell(row=i, column=1).value
        for j in range(2, i):
            j_pos = current_sheet.cell(row=1, column=j).value
            value_pos = current_sheet.cell(row=i, column=j).value
            fd.write(str(i_pos) + "\t" + str(j_pos) + "\t" + str(value_pos) + '\n')

    fd.close()

if __name__ == "__main__":
    file_path = "E:\\2019\\endfile\\t_cid.xlsx"
    name_txt = "t_cid"
    excel_to_txt(file_path, name_txt)