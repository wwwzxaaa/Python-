#coding=utf-8

from openpyxl import Workbook
def txt_to_excel(path_file):
    """
    将txt数据转存到excel数据中
    :param path_file: txt文件路径
    :return:
    """
    col_title = {}#存放列数据title
    col_title_list = []
    row_title = {}#存放行数据title
    row_title_list = []
    data_txt = []#存放txt数据
    col_num = 1#列的标号
    row_num = 1#行的标号
    with open(path_file, 'r', encoding='utf-8') as f:
        while True:
            lines = f.readline()  # 整行读取数据
            if lines:
                data_temp = [i for i in lines.split()]
                data_txt.append(data_temp)
                if row_title.get(data_temp[0]) == None:
                    row_title.update({data_temp[0] : row_num})
                    row_title_list.append(data_temp[0])
                    row_num += 1
                if col_title.get(data_temp[1]) == None:
                    col_title.update({data_temp[1] : col_num})
                    col_title_list.append(data_temp[1])
                    col_num += 1
            else:
                break
    if len(col_title) == 0 or len(row_title) == 0 or len(data_txt) == 0:
        print("数据存在问题")
    #生成excel数据
    Path_excel =  path_file[0:-4]+ ".xlsx"#定义excel文件名
    #保持文件
    ExcelFile = Workbook()
    sheet = ExcelFile.active
    sheet.cell(row=1, column=1, value="CID")
    #保存列title
    for i in range(1, row_num):
        sheet.cell(row=i+1, column=1, value=row_title_list[i-1])
    #保存行title
    for i in range(1, col_num):
        sheet.cell(row=1, column=i+1, value=col_title_list[i-1])
    #保存中间数据
    for item in data_txt:
        sheet.cell(row=row_title.get(item[0])+1, column=col_title.get(item[1])+1, value=item[2])

    ExcelFile.save(Path_excel)  # 保存工作簿





if __name__ == "__main__":
    file_path = "f-screen_target_value.txt"
    txt_to_excel(file_path)