#coding=utf-8

import openpyxl
def excel_to_Symmetric(path_file):
    """
    将excel中的数据转换为对称矩阵，对角线元素为1，其他对称位置如果冲突则取最大值
    :param path_file: excel文件路径
    :return:
    """
    wb = openpyxl.load_workbook(path_file)
    current_sheet = wb.active
    maxRow = current_sheet.max_row
    maxColumn = current_sheet.max_column
    if maxRow != maxColumn:
        print("数据存在问题:最大行数和最大列数的数值不相等")
        return
    #修正数据
    #对角位置修正为1
    for i in range(2, maxRow+1):
        current_sheet.cell(row=i, column=i, value=1)
    diff_num = 0
    #对称位置遍历
    for i in range(2, maxRow+1):
        for j in range(2, i):
            value_1 = float(current_sheet.cell(row=i, column=j).value)
            value_2 = float(current_sheet.cell(row=j, column=i).value)
            if value_1 != value_2:
                pos_row_1 = str(current_sheet.cell(row=i, column=1).value)
                pos_row_2 = str(current_sheet.cell(row=j, column=1).value)
                pos_col_1 = str(current_sheet.cell(row=1, column=j).value)
                pos_col_2 = str(current_sheet.cell(row=1, column=i).value)
                print("坐标位置(" + pos_row_1 + "," + pos_col_1 + ")和坐标位置（" + pos_row_2 + "," + pos_col_2 + ")数值存在冲突，分别为：" + str(value_1) + " 和 " + str(value_2))
                diff_num += 1
                value_3 = max(value_1, value_2)
                current_sheet.cell(row=i, column=j, value=value_3)
                current_sheet.cell(row=j, column=i, value=value_3)

    print("数据冲突的位置的组数为：" + str(diff_num))
    #保存excel数据
    Path_excel =  path_file[0:-5]+ "_Symmetric.xlsx"#定义excel文件名
    wb.save(Path_excel)  # 保存工作簿





if __name__ == "__main__":
    file_path = "end_simcomp_result.xlsx"
    excel_to_Symmetric(file_path)